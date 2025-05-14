#!/usr/bin/env python
from openpyxl import Workbook, load_workbook
import create_cyberark_accounts

# The script is used to modify the timeslot for OS patching


def main():
    timeslots = {
        1: "0000 - 0100",
        2: "0100 - 0200",
        3: "0200 - 0300",
        4: "0300 - 0400",
        5: "0400 - 0500",
        6: "0500 - 0600",
        7: "0600 - 0700",
        8: "0700 - 0800"
    }

    CATALIST='20250417 0000-0800.xlsx'
    hosts = []
    workbook = load_workbook(filename=CATALIST)
    sheet = workbook.active
    check_col = None
    letter_col = None
    count = 0
    for row in sheet.iter_rows():
        count += 1
        # Capture the hostnames
        hosts.append(row[0].value)
        for c in row:
            seq_col = str(c.value)
            seq_col = seq_col.lower()
            if "patching sequence" in seq_col:
                check_col = c.column
                letter_col = c.column_letter

    new_col = chr(ord(letter_col) + 1)
    cell_range = f"{letter_col}1:{letter_col}100"
    count = 1
    data = {}
    for column in sheet[cell_range]:
        for cell in column:
            if cell.value == 1:
                sheet[f"{new_col}{count}"] = timeslots[cell.value]
            elif cell.value == 2:
                sheet[f"{new_col}{count}"] = timeslots[cell.value]
            elif cell.value == 3:
                sheet[f"{new_col}{count}"] = timeslots[cell.value]
            elif cell.value == 4:
                sheet[f"{new_col}{count}"] = timeslots[cell.value]
            elif cell.value == 5:
                sheet[f"{new_col}{count}"] = timeslots[cell.value]
            elif cell.value == 6:
                sheet[f"{new_col}{count}"] = timeslots[cell.value]
            elif cell.value == 7:
                sheet[f"{new_col}{count}"] = timeslots[cell.value]
            elif cell.value == 8:
                sheet[f"{new_col}{count}"] = timeslots[cell.value]
            count += 1

    sheet.title = "Update Schedule"
    workbook.save(CATALIST)

    outputs = ""
    # Remove the header hostname
    hosts.pop(0)
    # Generate the cyberark accounts from hosts array
    # Create a string variable
    for h in hosts:
        if h:
            outputs += create_cyberark_accounts.get_account_name(h)

    # Append the default accounts with the variable above
    outputs += create_cyberark_accounts.DEFAULT_ACC

    # Create the data structure for excel sheet
    data = []
    accounts = outputs.split("\n")
    for a in accounts:
        data.append(a.split(","))

    workbook = load_workbook(filename=CATALIST)
    workbook.create_sheet(title="Cyberark")
    workbook.active = workbook["Cyberark"]
    sheet = workbook.active
    sheet.append(["hostname","account","team","description"])
    for row in data:
        sheet.append(row)

    workbook.save(CATALIST)

if __name__ == "__main__":
    main()
