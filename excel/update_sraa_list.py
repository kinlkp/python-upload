#!/usr/bin/env python


import datetime
from openpyxl import Workbook, load_workbook

# The script is used to modify the timeslot for OS patching

NAME="SRAA - Asset identification and valuation registry_with4.8_v2.xlsx"

workbook = load_workbook(filename=NAME)
sheet = workbook.active
max_row = sheet.max_row

def edit_sheet(hostname):
    found = 0
    today = datetime.date.today()
    for row in sheet.iter_rows():
        if row[2].value == hostname:
            if not row[8].value:
                row[7].value = today
                row[8].value = "Completed"
                found = 1
    if not found:
        sheet.append(["","",hostname, "", "", "", "", today, "Completed"])    

with open("s.list") as s:
    for h in s:
        edit_sheet(h.rstrip())

workbook.save(NAME)





