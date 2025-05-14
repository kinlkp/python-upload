#!/usr/bin/env python

import openpyxl
import sys
import re

from openpyxl import load_workbook

filename = sys.argv[1]
pattern = r'(.*)Should be\s+:\s*(.*)'

workbook = load_workbook(filename=filename)
sheet = workbook.active

all_rows = sheet.iter_rows()
name_col = ""
detail_col = ""
for row in all_rows:
    for c in row:
        if c.value == 'Name':
            name_col = c.column -1
        if c.value == 'Detail Path':
            detail_col = c.column -1
            break
    try:
        if 'RHEL' in row[name_col].value:
            # print(row[name_col].value)
            print(row[detail_col].value)
    except:
        print(row[2].value)

